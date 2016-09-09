import openpyxl
from openpyxl.styles import Border, Side
import csv
from shutil import copyfile
from pdb import set_trace

def TSV2xlsx(tsv_filename):
	xlsx_filename = tsv_filename + ".xlsx"
	workbook = openpyxl.Workbook()
	worksheet = workbook.active
	tsv_reader = csv.reader(open(tsv_filename, 'rb'), delimiter = '\t')
	for row in tsv_reader:
		for index, item in enumerate(row):
			try:
				int(item)
				row[index] = int(item)
			except ValueError:
				try:
					float(item)
					row[index] = float(item)
				except ValueError:
					continue
		worksheet.append(row)
	workbook.save(xlsx_filename)
	return xlsx_filename

def FormatTSV(tsv_filename, header = False):
	xlsx_filename = TSV2xlsx(tsv_filename)
	wb = openpyxl.load_workbook(xlsx_filename)
	sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	prev_name = ''
	for row in range(2, sheet.max_row+1):
		if sheet['A' + str(row)].value != prev_name:
			firstCell = sheet['A' + str(row)]
			lastCell = sheet.cell(row=row, column=sheet.max_column)
			firstCell.border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'))
			lastCell.border = Border(right=Side(border_style='thin'), top=Side(border_style='thin'))
			top_border = Border(top=Side(border_style='thin'))
			for col in range(2, sheet.max_column):
				sheet.cell(row=row, column=col).border = top_border
			prev_name = firstCell.value
		else:
			firstCell = sheet['A' + str(row)]
			lastCell = sheet.cell(row=row, column=sheet.max_column)
			firstCell.border = Border(left=Side(border_style='thin'))
			lastCell.border = Border(right=Side(border_style='thin'))
	sheet.cell(row=row, column = 1).border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'))
	sheet.cell(row=row, column = sheet.max_column).border = Border(right=Side(border_style='thin'), bottom=Side(border_style='thin'))
	for col in range(2, sheet.max_column):
		sheet.cell(row=row, column=col).border = Border(bottom=Side(border_style='thin'))
	if header:
		sheet.freeze_panes = 'A2'
	wb.save(xlsx_filename)

#FormatTSV('/Users/cdisuchicago/Documents/DENV_Minion_fasta/DENV2_blastn_DENV2NGS.xml.filtered_summary', header=True)
#FormatTSV('/Users/cdisuchicago/Documents/dengue_database/all_dengue_nucleotide.fasta.compiled.xml.summary.tsv', header = True)
FormatTSV('/Users/cdisuchicago/Documents/west_nile_database/all_west_nile_nucleotide.fasta.compiled.xml.summary.tsv', header = True)
FormatTSV('/Users/cdisuchicago/Documents/tick-borne_encephalitis_database/all_tick-borne_encephalitis_nucleotide.fasta.compiled.xml.summary.tsv', header = True)
FormatTSV('/Users/cdisuchicago/Documents/yellow_fever_database/all_yellow_fever_nucleotide.fasta.compiled.xml.summary.tsv', header = True)
